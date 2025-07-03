from typing import Generator

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from mireport.excelutil import CellRangeDimensions, getEffectiveCellRangeDimensions


@pytest.fixture
def sample_worksheet() -> Generator[Worksheet, None, None]:
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active

    # Fill sample data
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Data1"
    ws["B2"] = None
    ws["A3"] = None
    ws["B3"] = "Data2"
    ws["C1"] = None  # Intentionally blank

    yield ws


def test_filled_and_empty_cells(sample_worksheet: Worksheet) -> None:
    cr: CellRange = CellRange("A1:C3")
    dims: CellRangeDimensions = getEffectiveCellRangeDimensions(sample_worksheet, cr)

    assert dims.countPopulated == 4
    assert dims.width == 2  # Column C is completely empty
    assert dims.height == 3


def test_completely_empty_range(sample_worksheet: Worksheet) -> None:
    cr: CellRange = CellRange("D1:E5")
    dims: CellRangeDimensions = getEffectiveCellRangeDimensions(sample_worksheet, cr)

    assert dims.countPopulated == 0
    assert dims.width == 1  # min width enforced
    assert dims.height == 1  # min height enforced


def test_partially_filled_rows(sample_worksheet: Worksheet) -> None:
    sample_worksheet["C2"] = "ExtraData"
    cr: CellRange = CellRange("A1:C3")
    dims: CellRangeDimensions = getEffectiveCellRangeDimensions(sample_worksheet, cr)

    assert dims.countPopulated == 5
    assert dims.width == 3  # All columns have data now
    assert dims.height == 3


def test_single_cell_range(sample_worksheet: Worksheet) -> None:
    cr: CellRange = CellRange("A1")
    dims: CellRangeDimensions = getEffectiveCellRangeDimensions(sample_worksheet, cr)

    assert dims.countPopulated == 1
    assert dims.width == 1
    assert dims.height == 1
