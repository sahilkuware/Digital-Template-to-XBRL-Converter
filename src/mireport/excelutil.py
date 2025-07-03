import re
from datetime import date, datetime, time
from pathlib import Path
from typing import (
    BinaryIO,
    Iterator,
    Literal,
    NamedTuple,
    Optional,
    TypeAlias,
    Union,
    overload,
)

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, MergedCell, ReadOnlyCell
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from mireport.exceptions import OpenPyXlRelatedException

_CellType: TypeAlias = ReadOnlyCell | MergedCell | Cell
_CellValue: TypeAlias = bool | float | int | str | datetime | date | time | None

EXCEL_PLACEHOLDER_VALUE = "#VALUE!"


def checkExcelFilePath(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f'"{path}" is not a file.')
    elif path.suffix != ".xlsx":
        raise Exception(f'"{path}" is not a supported (.xlsx) Excel file')


def loadExcelFromPathOrFileLike(pathOrFile: Path | BinaryIO) -> Workbook:
    wb = load_workbook(
        filename=pathOrFile, read_only=False, data_only=True, rich_text=True
    )
    return wb


def excelCellRef(worksheet: Worksheet, cell: _CellType) -> str:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    ref = f"{quote_sheetname(worksheet.title)}!{absolute_coordinate(cell.coordinate)}"
    return ref


def excelCellRangeRef(worksheet: Worksheet, cellRange: CellRange) -> str:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    ref = f"{quote_sheetname(worksheet.title)}!{absolute_coordinate(cellRange.coord)}"
    return ref


def excelCellOrCellRangeRef(
    worksheet: Worksheet, cellRange: CellRange, cell: _CellType | None
) -> str:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    if cell is not None:
        return excelCellRef(worksheet, cell)
    elif cellRange is not None:
        return excelCellRangeRef(worksheet, cellRange)
    else:
        return None


def excelDefinedNameRef(
    definedName: Optional[DefinedName], cell: Optional[_CellType] = None
) -> Optional[str]:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    if definedName is None:
        return None

    destinations = list(definedName.destinations)
    match len(destinations):
        case 1:
            sheet_name, cell_range = destinations[0]
            if cell is not None:
                coord = cell.coordinate
            else:
                coord = cell_range
            ref = f"{quote_sheetname(sheet_name)}!{absolute_coordinate(coord)}"
            return ref
        case _:
            return None


def getNamedRanges(wb: Workbook) -> dict:
    data = {}
    for dn in list(wb.defined_names.values()):
        sheet_name, cell_range = list(dn.destinations)[0]
        if not cell_range:
            continue
        cr = CellRange(cell_range)
        if (
            cr.min_col is None
            or cr.min_row is None
            or cr.max_col is None
            or cr.max_row is None
        ):
            raise OpenPyXlRelatedException(
                f"Cell range bounds expected to be int but actually None {cr=}"
            )
        width: int = cr.max_col - cr.min_col
        height: int = cr.max_row - cr.min_row
        ws = wb[sheet_name]
        if not width and not height:
            cell = ws[cell_range]
            cells = [cell.value]
        else:
            cells = []
            for row in ws[cell_range]:
                cells.extend([c.value for c in row])
        data[dn.name] = cells
    return data


def get_decimal_places(cell: _CellType) -> int:
    """
    Returns the number of decimal places in the cell's number format.
    For example, a format of '0.00' would return 2.
    """
    number_format = cell.number_format

    # Match typical decimal number formats like '0.00', '#,##0.000', etc.
    match = re.search(r"\.(0+)", number_format)
    if match:
        return len(match.group(1))

    # Handle cases like percentage formats '0.0%' or '0.000%'
    match_percent = re.search(r"\.(0+)%", number_format)
    if match_percent:
        return len(match_percent.group(1))

    # Catch general cases like scientific notation '0.00E+00'
    match_sci = re.search(r"\.(0+)[eE]", number_format)
    if match_sci:
        return len(match_sci.group(1))

    return 0  # No decimal part found


@overload
def getCellRangeIterator(
    ws: Worksheet,
    cr: CellRange,
    row_start: Optional[int] = None,
    col_start: Optional[int] = None,
    group_by_row: Literal[False] = False,
) -> Iterator[tuple[int, int, _CellType]]: ...


@overload
def getCellRangeIterator(
    ws: Worksheet,
    cr: CellRange,
    row_start: Optional[int] = None,
    col_start: Optional[int] = None,
    group_by_row: Literal[True] = True,
) -> Iterator[tuple[int, tuple[_CellType, ...]]]: ...


def getCellRangeIterator(
    ws: Worksheet,
    cr: CellRange,
    row_start: Optional[int] = None,
    col_start: Optional[int] = None,
    group_by_row: bool = False,
) -> Iterator[Union[tuple[int, int, _CellType], tuple[int, tuple[_CellType, ...]]]]:
    """Iterates over cells in the given range, supporting both standard and row-grouped modes."""

    if cr.min_row is None or cr.min_col is None:
        raise OpenPyXlRelatedException(
            f"Cell range bounds expected to be int but actually None {cr=}"
        )
    actual_row_start: int = cr.min_row
    if row_start is not None:
        actual_row_start = row_start

    actual_col_start: int = cr.min_col
    if col_start is not None:
        actual_col_start = col_start

    for rnum, row in enumerate(
        ws.iter_rows(
            min_row=cr.min_row,
            min_col=cr.min_col,
            max_row=cr.max_row,
            max_col=cr.max_col,
        ),
        start=actual_row_start,
    ):
        if group_by_row:
            yield rnum, row  # Yield row number and tuple of cells
        else:
            for cnum, cell in enumerate(row, start=actual_col_start):
                yield rnum, cnum, cell  # Yield row number, column number, and cell


class CellRangeDimensions(NamedTuple):
    width: int
    height: int
    cellsAccessed: set[tuple[str, int, int]]
    cellsPopulated: set[tuple[str, int, int]]

    @property
    def countAccessed(self) -> int:
        return len(self.cellsAccessed)

    @property
    def countPopulated(self) -> int:
        return len(self.cellsPopulated)


def getEffectiveCellRangeDimensions(
    ws: Worksheet, cell_range: CellRange
) -> CellRangeDimensions:
    cols_not_empty: set[int] = set()
    cols_empty: set[int] = set()
    populated_rows: set[int] = set()
    populatedCellCount: set[tuple[str, int, int]] = set()
    cellCount: set[tuple[str, int, int]] = set()

    last_rnum = None
    empty_row = True
    sheetName = ws.title
    for rnum, cnum, cell in getCellRangeIterator(ws, cell_range):
        cellCount.add((sheetName, rnum, cnum))
        if last_rnum is None:
            last_rnum = rnum

        if rnum != last_rnum:
            if not empty_row:
                populated_rows.add(last_rnum)
            last_rnum = rnum
            empty_row = True

        if cell.value is not None:
            populatedCellCount.add((sheetName, rnum, cnum))
            empty_row = False
            cols_not_empty.add(cnum)
        else:
            cols_empty.add(cnum)
    else:
        if not empty_row:
            populated_rows.add(rnum)

    definitely_empty_cols = cols_empty - cols_not_empty
    total_cols = len(cols_not_empty.union(cols_empty))
    width = max(1, total_cols - len(definitely_empty_cols))
    height = max(1, len(populated_rows))
    return CellRangeDimensions(
        width=width,
        height=height,
        cellsAccessed=cellCount,
        cellsPopulated=populatedCellCount,
    )
