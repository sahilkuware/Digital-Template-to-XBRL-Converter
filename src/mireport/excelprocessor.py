import difflib
import logging
import re
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from itertools import combinations
from pathlib import Path
from typing import BinaryIO, NamedTuple, Optional
from xml.sax.saxutils import escape as xml_escape

from dateutil.parser import parse as parse_datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from mireport.conversionresults import (
    ConversionResultsBuilder,
    MessageType,
    Severity,
)
from mireport.excelutil import (
    EXCEL_PLACEHOLDER_VALUE,
    _CellType,
    _CellValue,
    excelCellOrCellRangeRef,
    excelCellRangeRef,
    excelCellRef,
    excelDefinedNameRef,
    get_decimal_places,
    getCellRangeIterator,
    getEffectiveCellRangeDimensions,
    loadExcelFromPathOrFileLike,
)
from mireport.exceptions import EarlyAbortException, InlineReportException
from mireport.taxonomy import (
    VSME_ENTRY_POINT,
    Concept,
    QName,
    Taxonomy,
    getTaxonomy,
    listTaxonomies,
)
from mireport.xbrlreport import FactBuilder, InlineReport, _FactValue

L = logging.getLogger(__name__)


def _loadVsmeDefaults(bits: dict) -> None:
    VSME_DEFAULTS.update(bits)


EE_SET_DESIRED_EMPTY_PLACEHOLDER_VALUE = "None"

EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE = ("-", EXCEL_PLACEHOLDER_VALUE)

VSME_DEFAULTS: dict = dict()


def cleanUnitTextFromExcel(unitTest: str, replacements: dict[str, str]) -> str:
    new = unitTest
    for original, replacement in replacements.items():
        new = new.replace(original, replacement)
    return new


def eeDomainAsText(concept: Concept) -> str:
    domain = concept.getEEDomain()
    if not domain:
        return "(none)"
    out = [f'"{a.getStandardLabel()}" [{a.qname}]' for a in concept.getEEDomain()]
    return ", ".join(out)


def conceptsToText(concepts: Iterable[Concept]) -> str:
    return ", ".join(sorted(str(c.qname) for c in concepts))


class ComplexUnit(NamedTuple):
    numerator: list[QName]
    denominator: list[QName]


@dataclass(slots=True, eq=True, frozen=True)
class CellRangeMetadata:
    definedName: DefinedName
    worksheet: Worksheet
    cellRange: CellRange
    effectiveWidth: int
    effectiveHeight: int
    cellsPopulated: int


@dataclass(slots=True, eq=True, frozen=True)
class CellAndXBRLMetadataHolder(CellRangeMetadata):
    concept: Concept

    @classmethod
    def fromCellRangeMetadata(
        cls, holder: CellRangeMetadata, concept: Concept
    ) -> "CellAndXBRLMetadataHolder":
        """
        Create a CellAndXBRLMetadataHolder instance from a CellRangeMetadataHolder and a Concept.

        :param holder: A CellRangeMetadataHolder instance
        :param concept: A Concept instance to be associated with the new CellAndXBRLMetadataHolder
        :return: A new instance of CellAndXBRLMetadataHolder
        """
        return cls(
            definedName=holder.definedName,
            worksheet=holder.worksheet,
            cellRange=holder.cellRange,
            effectiveWidth=holder.effectiveWidth,
            effectiveHeight=holder.effectiveHeight,
            concept=concept,
            cellsPopulated=holder.cellsPopulated,
        )


class TableXBRLContents(NamedTuple):
    primaryItems: list[CellAndXBRLMetadataHolder]
    explicitDimensions: list[CellAndXBRLMetadataHolder]
    typedDimensions: list[CellAndXBRLMetadataHolder]
    units: list[CellAndXBRLMetadataHolder]


class ExcelProcessor:
    def __init__(
        self,
        excelPathOrFileLike: Path | BinaryIO,
        results: ConversionResultsBuilder,
        defaults: dict,
        /,
    ):
        self._results = results
        self._defaults = defaults
        self._excelPathOrFileLike: Path | BinaryIO = excelPathOrFileLike

        # Populated from config file
        self._configDataTypeToUnitMap: dict[QName, QName] = {}
        self._configUnitIdsToMeasures: dict[str, ComplexUnit] = {}
        self._configCellValuesToTaxonomyLabels: dict[str, str] = {}
        self._configConceptToUnitMap: dict[Concept, QName] = {}
        self._configCellUnitReplacements: dict[str, str] = {}

        # Populated from Excel sheet
        self._unusedDefinedNames: set[DefinedName] = set()
        self._conceptToUnitHolderMap: dict[Concept, CellAndXBRLMetadataHolder] = {}
        self._definedNameToXBRLMap: dict[DefinedName, CellAndXBRLMetadataHolder] = {}
        self._presetDimensions: dict[
            CellAndXBRLMetadataHolder, dict[Concept, Concept]
        ] = defaultdict(dict)
        self._tableRelatedNames: dict[CellAndXBRLMetadataHolder, TableXBRLContents] = {}

        # Not yet initialised. Need setting early
        self._workbook: Workbook
        self._report: InlineReport

    @property
    def taxonomy(self) -> Taxonomy:
        return self._report.taxonomy

    @property
    def unusedNames(self) -> list[str]:
        return sorted(dn.name for dn in self._unusedDefinedNames if dn.name)

    def populateReport(self) -> InlineReport:
        """
        Add facts to InlineReport from the provided Excel workbook.
        The workbook is close()d before this method returns
        """
        try:
            self._loadWorkbook()
            assert self._workbook

            self._verifyEntryPoint()
            self.abortEarlyIfErrors()
            assert self._report

            self.getAndValidateRequiredMetadata()
            self._processConfiguration()
            self.abortEarlyIfErrors()

            self._recordNamedRanges()
            self._processNamedRanges()
            self._processNamedRangeTables()
            self._createNamedPeriods()
            self.createSimpleFacts()
            self.createTableFacts()
            self.checkForUnhandledItems()
            return self._report
        except EarlyAbortException as eae:
            self._results.addMessage(
                f"Excel conversion aborted early. {eae}",
                Severity.ERROR,
                MessageType.ExcelParsing,
            )
            raise
        except Exception as e:
            self._results.addMessage(
                f"Exception encountered during processing. {e}",
                Severity.ERROR,
                MessageType.ExcelParsing,
            )
            L.exception("Exception encountered", exc_info=e)
            raise
        finally:
            self._workbook.close()

    def _loadWorkbook(self) -> None:
        self._workbook = loadExcelFromPathOrFileLike(self._excelPathOrFileLike)

    def _recordNamedRanges(self) -> None:
        self._unusedDefinedNames.update(
            dn
            for dn in self._workbook.defined_names.values()
            if dn.name and not dn.name.startswith(("enum_", "template_"))
        )

    def getDefinedNameForString(self, name: str) -> Optional[DefinedName]:
        """
        Get the DefinedName for a given name string or None if it is not present.
        """
        if name in self._workbook.defined_names:
            return self._workbook.defined_names[name]
        return None

    def _verifyEntryPoint(self) -> None:
        name = self._defaults.get("entryPoint", "")
        entryPoint = self.getSingleStringValue(name)
        validEntryPoints = set(listTaxonomies())
        if not entryPoint and VSME_ENTRY_POINT in validEntryPoints:
            self._results.addMessage(
                "Excel template does not specify taxonomy entry point. Please use an official template.",
                Severity.ERROR,
                MessageType.ExcelParsing,
                excel_reference=excelDefinedNameRef(self.getDefinedNameForString(name)),
            )
        elif entryPoint not in validEntryPoints:
            self._results.addMessage(
                f"Excel report is for an unsupported taxonomy. Excel wants: {entryPoint=}. We support: {sorted(validEntryPoints)}",
                Severity.ERROR,
                MessageType.ExcelParsing,
                excel_reference=excelDefinedNameRef(self.getDefinedNameForString(name)),
            )

        self.abortEarlyIfErrors()
        taxonomy = getTaxonomy(entryPoint)
        self._report = InlineReport(taxonomy)
        self._report.addSchemaRef(entryPoint)

    def getAndValidateRequiredMetadata(self) -> None:
        defaults = self._defaults
        entityIdentifierSchemeLabelToURIs: dict[str, str] = {
            k: v for k, v in defaults["entityIdentifierLabelsToSchemes"].items()
        }
        if "aoix" in defaults:
            for aoixName, namedRangeName in defaults["aoix"].items():
                if namedRangeName not in self._workbook.defined_names:
                    self._results.addMessage(
                        f"Excel report must have a value for named range {namedRangeName}.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                    )
                    continue
                if aoixName == "entity-scheme":
                    lookup_key = (
                        self.getSingleStringValue(namedRangeName)
                        .strip()
                        .replace(" ", "")
                        .lower()
                    )
                    aoixValue = entityIdentifierSchemeLabelToURIs.get(lookup_key)
                else:
                    aoixValue = self.getSingleStringValue(namedRangeName).strip()

                if (
                    not aoixValue
                    or aoixValue in EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE
                ):
                    self._results.addMessage(
                        f"Excel report must have a valid value for named range {namedRangeName}.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        excel_reference=excelDefinedNameRef(
                            self.getDefinedNameForString(namedRangeName)
                        ),
                    )
                    continue
                self._report.setDefaultAspect(aoixName, aoixValue)

        if "periods" in defaults:
            for period in defaults["periods"]:
                failed = False

                try:
                    startName = period["start"]
                    startDate = self.getSingleDateValue(startName)
                except Exception as e:
                    self._results.addMessage(
                        f"Excel report must have a valid date for named range {startName}. Exception: {e}",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        excel_reference=excelDefinedNameRef(
                            self.getDefinedNameForString(startName)
                        ),
                    )
                    failed = True

                try:
                    endName = period["end"]
                    endDate = self.getSingleDateValue(endName)
                except Exception as e:
                    self._results.addMessage(
                        f"Excel report must have a valid date for named range {endName}. Exception: {e}",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        excel_reference=excelDefinedNameRef(
                            self.getDefinedNameForString(endName)
                        ),
                    )
                    failed = True

                if not failed and startDate > endDate:
                    self._results.addMessage(
                        f"Start date {startDate} is after end date {endDate}.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        excel_reference=excelDefinedNameRef(
                            self.getDefinedNameForString(period["start"])
                        ),
                    )
                    failed = True

                name = period["name"]
                if not failed and self._report.addDurationPeriod(
                    name,
                    startDate,
                    endDate,
                ):
                    self._report.setDefaultPeriodName(name)

        if "report" in defaults:
            entityName_namedRange = defaults["report"]["entity-name"]
            if entityName_namedRange in self._workbook.defined_names:
                self._report.setEntityName(
                    self.getSingleStringValue(entityName_namedRange)
                )
            else:
                self._results.addMessage(
                    f"Excel report must have a value for named range {entityName_namedRange}.",
                    Severity.ERROR,
                    MessageType.ExcelParsing,
                )

    def abortEarlyIfErrors(self) -> None:
        if self._results.hasErrors():
            raise EarlyAbortException(
                "Excel report is missing required named ranges or data. Please check the report and try again."
            )

    def _processConfiguration(self) -> None:
        defaults = self._defaults
        if "dataTypesToUnits" in defaults:
            for dataType, unitType in defaults["dataTypesToUnits"].items():
                self._configDataTypeToUnitMap[
                    self.taxonomy.QNameMaker.fromString(dataType)
                ] = self.taxonomy.QNameMaker.fromString(unitType)

        if "unitIdsToMeasures" in defaults:
            for unitId, unitDict in defaults["unitIdsToMeasures"].items():
                numerators: list[QName] = [
                    qname
                    for m in unitDict.get("numerator", [])
                    if (qname := self.taxonomy.UTR.getQNameForUnitId(m)) is not None
                ]
                denominators: list[QName] = [
                    qname
                    for m in unitDict.get("denominator", [])
                    if (qname := self.taxonomy.UTR.getQNameForUnitId(m)) is not None
                ]
                self._configUnitIdsToMeasures[unitId] = ComplexUnit(
                    numerator=numerators, denominator=denominators
                )

        if "conceptsToUnits" in defaults:
            for conceptQname, unitQname in defaults["conceptsToUnits"].items():
                self._configConceptToUnitMap[self.taxonomy.getConcept(conceptQname)] = (
                    self.taxonomy.QNameMaker.fromString(unitQname)
                )

        if "cellValuesToTaxonomyLabels" in defaults:
            self._configCellValuesToTaxonomyLabels.update(
                defaults["cellValuesToTaxonomyLabels"]
            )

        if "cellUnitReplacements" in defaults:
            self._configCellUnitReplacements.update(defaults["cellUnitReplacements"])

    def getSingleCell(
        self,
        definedName: DefinedName | str | CellAndXBRLMetadataHolder | CellRangeMetadata,
        *,
        row: int = -1,
        column: int = -1,
    ) -> Optional[_CellType]:
        if isinstance(definedName, str):
            definedName = self._workbook.defined_names.get(definedName)
            if definedName is None:
                return None

        stuff: CellAndXBRLMetadataHolder | CellRangeMetadata
        if isinstance(definedName, DefinedName):
            found = self._definedNameToXBRLMap.get(definedName)
            if found is None:
                if (crh := self._getCellRange(definedName)) is not None:
                    self._unusedDefinedNames.discard(definedName)
                    stuff = crh
                else:
                    return None
            else:
                stuff = found
        elif isinstance(definedName, (CellAndXBRLMetadataHolder, CellRangeMetadata)):
            stuff = definedName

        self._unusedDefinedNames.discard(stuff.definedName)

        cr = stuff.cellRange
        ws = stuff.worksheet

        if not all(
            x is not None for x in (cr.min_row, cr.max_row, cr.min_col, cr.max_col)
        ):
            self._results.addMessage(
                f"Named range {stuff.definedName.name} has an invalid cell range {cr.bounds}.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelDefinedNameRef(stuff.definedName),
            )
            return None
        assert (
            cr.min_row is not None
            and cr.max_row is not None
            and cr.min_col is not None
            and cr.max_col is not None
        )

        if cr.min_row == cr.max_row:
            # Special case: single cell range, disregard and passed in row or
            # column and use the range's min_row
            row = cr.min_row

        if cr.min_col == cr.max_col:
            # Special case: single cell range, disregard and passed in row or
            # column and use the range's min_col
            column = cr.min_col

        if row == -1:
            row = cr.min_row
        if column == -1:
            column = cr.min_col

        if not (cr.min_row <= row <= cr.max_row):
            self._results.addMessage(
                f"Row {row} has not been specified correctly.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelCellRangeRef(ws, cr),
            )
            row = cr.min_row
        if not (cr.min_col <= column <= cr.max_col):
            self._results.addMessage(
                f"Column {column} has not been specified correctly.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelCellRangeRef(ws, cr),
            )
            column = cr.min_col

        rows = list(
            ws.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column)
        )
        match len(rows):
            case 0:
                return None
            case 1:
                cells = rows[0]
            case _:
                return None

        match len(cells):
            case 0:
                cell = None
                self._results.addMessage(
                    "No cells found in row of this named range.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(ws, cr),
                )
            case 1:
                cell = cells[0]
            case _:
                cell = None
                self._results.addMessage(
                    f"More than one cell found in range but only expected one cell. {cells}",
                    Severity.ERROR,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(ws, cr),
                )

        if cell is None or cell.value is None:
            return None

        if cell.value == EXCEL_PLACEHOLDER_VALUE:
            self._results.addMessage(
                f"Excel cell has an invalid stored value {EXCEL_PLACEHOLDER_VALUE}. Please check the Excel formula for this specific cell.",
                Severity.ERROR,
                MessageType.ExcelParsing,
                excel_reference=excelCellRef(ws, cell),
            )
            return None
        return cell

    def getSingleValue(
        self,
        definedName: DefinedName | str,
        *,
        row: int = -1,
        column: int = -1,
    ) -> _CellValue:
        if (
            cell := self.getSingleCell(definedName, row=row, column=column)
        ) is not None:
            return cell.value
        return None

    def getSingleStringValue(
        self,
        definedName: DefinedName | str,
        *,
        row: int = -1,
        column: int = -1,
    ) -> str:
        value = self.getSingleValue(definedName, row=row, column=column)
        return str(value) if value is not None else ""

    def getSimpleUnit(
        self, unitHolder: CellAndXBRLMetadataHolder, cell: _CellType
    ) -> Optional[QName]:
        if not cell.value:
            return None
        cellValue = str(cell.value).strip()
        candidates = [cellValue]
        candidates.extend(re.findall(r"\((.*?)\)", cellValue))
        possible_units = [
            unit
            for c in candidates
            if (unit := self.taxonomy.UTR.getQNameForUnitId(c)) is not None
        ]
        if not possible_units:
            candidates = [
                cleanUnitTextFromExcel(c, self._configCellUnitReplacements)
                for c in candidates
            ]
            possible_units = [
                unit
                for c in candidates
                if (unit := self.taxonomy.UTR.getQNameForUnitId(c)) is not None
            ]
            if possible_units:
                # our cleanUnitTextFromExcel fix up worked
                self._results.addMessage(
                    f"Workaround performed for mislabelled unit for {unitHolder.concept.qname}. Cell value '{cellValue}'. Unit ids now guessed '{possible_units}'",
                    Severity.WARNING,
                    MessageType.DevInfo,
                    taxonomy_concept=unitHolder.concept,
                    excel_reference=excelCellRef(unitHolder.worksheet, cell),
                )
        match len(possible_units):
            case 1:
                return possible_units[0]
            case 0:
                return None
            case _:
                self._results.addMessage(
                    f"Ambiguous unit specified in cell '{cellValue}'. Identified possible units: {possible_units}",
                    Severity.ERROR,
                    MessageType.ExcelParsing,
                    excel_reference=excelCellRef(unitHolder.worksheet, cell),
                )
                return None

    def _processNamedRanges(self) -> None:
        for dn in sorted(self._unusedDefinedNames, key=lambda d: d.name):
            concept = self.taxonomy.getConceptForName(dn.name)

            # TODO FIXME Temporary fix for the VSME taxonomy
            if dn.name == "IdentifierOfSitesInBiodiversitySensitiveAreasTypedAxis":
                concept = self.taxonomy.getConceptForName("IdentifierOfSiteTypedAxis")
            # TODO FIXME Temporary fix for the VSME taxonomy

            if concept is not None:
                if (crh := self._getCellRange(dn)) is not None:
                    self._definedNameToXBRLMap[dn] = (
                        CellAndXBRLMetadataHolder.fromCellRangeMetadata(
                            crh, concept=concept
                        )
                    )
            elif "_" in dn.name:
                conceptName, _, memberName = dn.name.partition("_")
                if "unit" == memberName:
                    if (
                        concept := self._report.taxonomy.getConceptForName(conceptName)
                    ) is not None and (crh := self._getCellRange(dn)) is not None:
                        self._conceptToUnitHolderMap[concept] = (
                            CellAndXBRLMetadataHolder.fromCellRangeMetadata(
                                crh, concept
                            )
                        )
                        self._unusedDefinedNames.remove(dn)
                else:
                    concept = self._report.taxonomy.getConceptForName(conceptName)
                    dimValue = self._report.taxonomy.getConceptForName(memberName)
                    crh = self._getCellRange(dn)
                    if crh is not None and concept is not None and dimValue is not None:
                        b = CellAndXBRLMetadataHolder.fromCellRangeMetadata(
                            crh, concept=concept
                        )
                        if (
                            dim
                            := self._report.taxonomy.getExplicitDimensionForDomainMember(
                                concept, dimValue
                            )
                        ) is not None:
                            self._definedNameToXBRLMap[dn] = b
                            self._presetDimensions[b][dim] = dimValue
                        else:
                            self._results.addMessage(
                                f"Domain member qualification set in named range {dn.name} but no dimension can be found for member.",
                                Severity.ERROR,
                                MessageType.DevInfo,
                            )
            if dn in self._definedNameToXBRLMap:
                self._unusedDefinedNames.remove(dn)
        self._results.addMessage(
            f"Excel file parsed ({self._results.numCellsPopulated} cells had data, with {self._results.numCellQueries} cells accessed).",
            Severity.INFO,
            MessageType.ExcelParsing,
        )

    def _getCellRange(self, dn: DefinedName) -> Optional[CellRangeMetadata]:
        all_destinations = list(dn.destinations)
        match len(all_destinations):
            case 0:
                self._results.addMessage(
                    f"Named range {dn.name} has no destinations specified. Ignoring.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                return None
            case 1:
                pass
            case _:
                self._results.addMessage(
                    f"Table {dn.name} has multiple destinations. Ignoring table.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                return None
        sheetName, cell_range = all_destinations[0]
        if not sheetName or not cell_range:
            self._results.addMessage(
                f"Named range {dn.name} has damaged cell reference {sheetName=} {cell_range=}",
                Severity.ERROR,
                MessageType.ExcelParsing,
            )
            return None
        try:
            ws = self._workbook[sheetName]
            cr = CellRange(cell_range)
        except Exception as e:
            L.exception("OpenPyXL is sad.", exc_info=e)
            return None
        dims = getEffectiveCellRangeDimensions(ws, cr)
        self._results.addCellQueries(dims.cellsAccessed)
        self._results.addCellsWithData(dims.cellsPopulated)
        return CellRangeMetadata(
            dn,
            ws,
            cr,
            effectiveHeight=dims.height,
            effectiveWidth=dims.width,
            cellsPopulated=len(dims.cellsPopulated),
        )

    def _processNamedRangeTables(self) -> None:
        # tables have one big named range A for the hypercube and then
        # additional named ranges B for each primary item and dimension within
        # it. Bs should always be wholly within A and Bs should never overlap
        # with each other.
        tables = [
            (dn, stuff)
            for dn, stuff in self._definedNameToXBRLMap.items()
            if stuff.concept in self.taxonomy.hypercubes
        ]
        concepts_in_excel = frozenset(
            stuff.concept for stuff in self._definedNameToXBRLMap.values()
        )

        hc_concepts_in_excel = frozenset(
            concept for concept in concepts_in_excel if concept.isHypercube
        )
        used_empty_hypercubes = self.taxonomy.emptyHypercubes.intersection(
            hc_concepts_in_excel
        )
        if used_empty_hypercubes:
            self._results.addMessage(
                # Someone forgot to put the hypercubes in the right relationships in the definition linkbase
                f"The following hypercubes exist and have corresponding named ranges but they cannot be used due to missing taxonomy definitions: {conceptsToText(used_empty_hypercubes)}.",
                Severity.ERROR,
                MessageType.DevInfo,
            )

        for table, table_stuff in tables:
            tableCr = table_stuff.cellRange
            tableWorksheet = table_stuff.worksheet
            table_concept = table_stuff.concept

            allPermittedConceptsForTable = self.taxonomy.getDimensionsForHypercube(
                table_concept
            ).union(
                {
                    concept
                    for concept in self.taxonomy.getPrimaryItemsForHypercube(
                        table_concept
                    )
                    if concept.isReportable or concept.isDimension
                }
            )
            missing_from_excel = allPermittedConceptsForTable.difference(
                concepts_in_excel
            )
            if missing_from_excel:
                self._results.addMessage(
                    f"Expected Dimensions or Primary Items for hypercube {table.name} have not been found: {conceptsToText(missing_from_excel)}.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                )

            candidates: list[CellAndXBRLMetadataHolder] = []
            extras_in_excel: set[CellAndXBRLMetadataHolder] = set()
            for dn, stuff in self._definedNameToXBRLMap.items():
                if tableWorksheet is not stuff.worksheet:
                    continue
                concept = stuff.concept
                if not (concept.isReportable or concept.isDimension):
                    continue
                if tableCr.issuperset(stuff.cellRange):
                    if concept in allPermittedConceptsForTable:
                        candidates.append(stuff)
                    else:
                        extras_in_excel.add(stuff)
                elif not tableCr.isdisjoint(stuff.cellRange):
                    extras_in_excel.add(stuff)

            if extras_in_excel:
                self._results.addMessage(
                    f"Extra named ranges found within/overlapping bounds of {table.name} named range but not supported by Hypercube {table_stuff.concept.qname}: {extras_in_excel}.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                )

            fishy = False
            for c1, c2 in combinations(candidates, 2):
                disjoint = c1.cellRange.isdisjoint(c2.cellRange)
                # same only makes sense for primary items, not for dimensions
                same = (
                    c1.concept.isReportable
                    and c2.concept.isReportable
                    and (c1.cellRange.bounds == c2.cellRange.bounds)
                )
                if not (disjoint or same):
                    fishy = True
                    self._results.addMessage(
                        f"Named range (table) {table.name} has named ranges (primary items or dimensions) {c1.definedName.name} and {c2.definedName.name} that are neither the same nor disjoint. Ignoring table.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                    )
                    break
            if not fishy:
                pItems = [c for c in candidates if c.concept.isReportable]
                eDims = [c for c in candidates if c.concept.isExplicitDimension]
                tDims = [c for c in candidates if c.concept.isTypedDimension]
                units = [
                    u
                    for p in pItems
                    if (u := self._conceptToUnitHolderMap.get(p.concept)) is not None
                ]
                self._tableRelatedNames[table_stuff] = TableXBRLContents(
                    primaryItems=pItems,
                    explicitDimensions=eDims,
                    typedDimensions=tDims,
                    units=units,
                )

        # Anything we're handling as a table should not be left in the list of standalone facts.
        for tableStuff, table_contents in self._tableRelatedNames.items():
            self._definedNameToXBRLMap.pop(tableStuff.definedName)
            tableDict: dict[str, list[CellAndXBRLMetadataHolder]] = (
                table_contents._asdict()
            )
            for name, partList in tableDict.items():
                for holder in partList:
                    if "units" == name:
                        self._conceptToUnitHolderMap.pop(holder.concept)
                    else:
                        self._definedNameToXBRLMap.pop(holder.definedName)

    def _createNamedPeriods(self) -> None:
        potentialPeriodHolders = [
            holder
            for holder in self._definedNameToXBRLMap.values()
            if holder.concept.isAbstract
        ]
        membersWithPotentialPeriods = {
            dimValue
            for dimPair in self._presetDimensions.values()
            for dimValue in dimPair.values()
        }
        periodHolders = [
            p
            for p in potentialPeriodHolders
            if p.concept in membersWithPotentialPeriods
        ]
        for periodHolder in periodHolders:
            dimValueDN = periodHolder.definedName
            namedPeriod = dimValueDN.name
            year = self.getSingleValue(dimValueDN)
            if year is None or year in EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE:
                self._definedNameToXBRLMap.pop(dimValueDN)
                continue

            if isinstance(year, bool) or not isinstance(year, float | int | str):
                self._results.addMessage(
                    f"Unable to extract year for {dimValueDN.name}. Cell value '{year}'",
                    Severity.ERROR,
                    MessageType.ExcelParsing,
                    taxonomy_concept=periodHolder.concept,
                    excel_reference=excelCellRangeRef(
                        periodHolder.worksheet, periodHolder.cellRange
                    ),
                )
                self._definedNameToXBRLMap.pop(dimValueDN)
                continue

            try:
                yearInt = int(year)
                self.getOrAddNamedPeriodForYear(namedPeriod, yearInt)
                self._definedNameToXBRLMap.pop(dimValueDN)
            except ValueError:
                self._results.addMessage(
                    f"Unable to convert value '{year}' to an integer.",
                    Severity.ERROR,
                    MessageType.ExcelParsing,
                    taxonomy_concept=periodHolder.concept,
                    excel_reference=excelCellRangeRef(
                        periodHolder.worksheet, periodHolder.cellRange
                    ),
                )
        return

    def createTableFacts(self) -> None:
        for tableStuff, table_contents in self._tableRelatedNames.items():
            tableDn = tableStuff.definedName
            primary_items = table_contents.primaryItems
            explicit_dimensions = table_contents.explicitDimensions
            typed_dimensions = table_contents.typedDimensions
            if 0 == len(primary_items):
                self._results.addMessage(
                    f"Table {tableDn.name} has no primary items defined. Skipping.",
                    Severity.ERROR,
                    MessageType.ExcelParsing,
                    excel_reference=excelCellRangeRef(
                        tableStuff.worksheet, tableStuff.cellRange
                    ),
                )
                continue

            for priItem in primary_items:
                concept = priItem.concept
                broken = False
                for rnum, row in getCellRangeIterator(
                    priItem.worksheet, priItem.cellRange, group_by_row=True
                ):
                    cells = [cell for cell in row if cell.value is not None]
                    match len(cells):
                        case 0:
                            cell = None
                            value = None
                        case 1:
                            cell = cells[0]
                            value = cell.value
                        case _:
                            # multiple columns is fine if it is a merged cell
                            # but multiple values is not unless we are an
                            # enumeration set
                            values = [c.value for c in cells]
                            cell = cells[0]
                            if concept.isEnumerationSet:
                                value = " ".join(values)
                            else:
                                self._results.addMessage(
                                    f"Primary item {priItem.definedName.name} spans multiple columns and has multiple values ({values}). Skipping.",
                                    Severity.ERROR,
                                    MessageType.ExcelParsing,
                                    excel_reference=excelCellOrCellRangeRef(
                                        priItem.worksheet, priItem.cellRange, cell
                                    ),
                                )
                                broken = True
                                break
                    if (
                        value is not None
                        and value is not False
                        and value not in EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE
                    ):
                        factBuilder = self._report.getFactBuilder()
                        factBuilder.setValue(value).setConcept(concept)

                        if (
                            presetDimensions := self._presetDimensions.get(priItem)
                        ) is not None:
                            for dim, dimValue in presetDimensions.items():
                                if (
                                    defaultValue := self.taxonomy.getDimensionDefault(
                                        dim
                                    )
                                ) is not None and dimValue != defaultValue:
                                    factBuilder.setExplicitDimension(dim, dimValue)

                        if concept.isNumeric:
                            unitHolder = None
                            sharedRange = False
                            for candidate in table_contents.units:
                                if candidate.concept == concept:
                                    unitHolder = candidate
                                    break

                            if unitHolder:
                                others = list(table_contents.units)
                                others.remove(unitHolder)
                                for candidate in others:
                                    if unitHolder.cellRange == candidate.cellRange:
                                        sharedRange = True

                            self.processNumeric(priItem, cell, factBuilder, value)
                            if not self.setUnitForName(
                                priItem,
                                factBuilder,
                                row=rnum,
                                specifiedUnitHolder=unitHolder,
                                sharedRange=sharedRange,
                            ):
                                continue

                        for td in typed_dimensions:
                            tdConcept = td.concept
                            tdCell = self.getSingleCell(td, row=rnum)
                            tdValue = None
                            if tdCell:
                                tdValue = tdCell.value
                            else:
                                L.info(f"{td.cellRange.bounds=}, {rnum=}")
                            if tdValue is not None:
                                if not isinstance(tdValue, _FactValue):
                                    tdValue = str(tdValue)
                                factBuilder.setTypedDimension(tdConcept, tdValue)
                            else:
                                broken = True
                                self._results.addMessage(
                                    f"Required typed dimension {tdConcept.qname} not set",
                                    Severity.ERROR,
                                    MessageType.Conversion,
                                    excel_reference=excelCellOrCellRangeRef(
                                        td.worksheet, td.cellRange, tdCell
                                    ),
                                )

                        for ed in explicit_dimensions:
                            edConcept = ed.concept
                            edCell = self.getSingleCell(ed, row=rnum)
                            edValue = None
                            if edCell:
                                edValue = edCell.value
                            else:
                                L.warning(
                                    f"Trying to access cell in named range {ed.definedName.name} {ed.cellRange.bounds=}, {rnum=}"
                                )
                                continue

                            if edValue is None:
                                self._results.addMessage(
                                    f"Required explicit dimension {edConcept.qname} not set. Cell value '{edValue}'",
                                    Severity.ERROR,
                                    MessageType.Conversion,
                                    excel_reference=excelCellOrCellRangeRef(
                                        ed.worksheet, ed.cellRange, edCell
                                    ),
                                )
                                broken = True
                                continue

                            memberConcept = self.taxonomy.getConceptForLabel(edValue)
                            if (
                                memberConcept is None
                                and (
                                    fake_value
                                    := self._configCellValuesToTaxonomyLabels.get(
                                        edValue
                                    )
                                )
                                is not None
                            ):
                                memberConcept = (
                                    self._report.taxonomy.getConceptForLabel(fake_value)
                                )

                            if memberConcept is not None:
                                factBuilder.setExplicitDimension(
                                    edConcept, memberConcept
                                )
                            else:
                                broken = True
                                self._results.addMessage(
                                    f"Required explicit dimension {edConcept.qname} not set. Cell value '{edValue}'",
                                    Severity.ERROR,
                                    MessageType.Conversion,
                                    excel_reference=excelCellOrCellRangeRef(
                                        ed.worksheet, ed.cellRange, edCell
                                    ),
                                )

                        if concept.isEnumerationSingle:
                            if (
                                eeValue := self._report.taxonomy.getConceptForLabel(
                                    value
                                )
                            ) is not None:
                                factBuilder.setHiddenValue(eeValue.expandedName)
                            else:
                                broken = True
                                self._results.addMessage(
                                    f"Unable to find EE concept for cell value '{value}'",
                                    Severity.ERROR,
                                    MessageType.Conversion,
                                    excel_reference=excelCellRef(
                                        priItem.worksheet, cell
                                    ),
                                )
                        elif concept.isEnumerationSet:
                            eeValues = []
                            for v in values:
                                if (
                                    eeValue := self._report.taxonomy.getConceptForLabel(
                                        v
                                    )
                                ) is not None:
                                    eeValues.append(eeValue)
                                else:
                                    broken = True
                                    self._results.addMessage(
                                        f"Unable to find EE concept for cell value '{value}'",
                                        Severity.ERROR,
                                        MessageType.Conversion,
                                    )
                            factBuilder.setHiddenValue(
                                " ".join(sorted(set(e.expandedName for e in eeValues)))
                            )
                        if broken:
                            if cell is not None:
                                ref = excelCellRef(priItem.worksheet, cell)
                            else:
                                ref = excelCellRangeRef(
                                    priItem.worksheet, priItem.cellRange
                                )

                            self._results.addMessage(
                                f"Unable to add fact with value '{value}'",
                                Severity.WARNING,
                                MessageType.Conversion,
                                excel_reference=ref,
                            )
                            continue
                        else:
                            self.addFactToReport(factBuilder, priItem)

    def addFactToReport(
        self, factBuilder: FactBuilder, holder: CellAndXBRLMetadataHolder
    ) -> bool:
        try:
            self._report.addFact(factBuilder.buildFact())
            return True
        except InlineReportException as i:
            self._results.addMessage(
                f"Unable to add fact. Encountered error: {i}",
                Severity.WARNING,
                MessageType.Conversion,
                excel_reference=excelCellRangeRef(holder.worksheet, holder.cellRange),
            )
        return False

    def setUnitForName(
        self,
        conceptHolder: CellAndXBRLMetadataHolder,
        factBuilder: FactBuilder,
        *,
        row: int = -1,
        specifiedUnitHolder: Optional[CellAndXBRLMetadataHolder] = None,
        sharedRange: Optional[bool] = None,
    ) -> bool:
        concept = conceptHolder.concept
        # See if we have a {conceptName}_unit named range with a cell value that
        # is a valid unit.
        unitHolder: Optional[CellAndXBRLMetadataHolder]
        if specifiedUnitHolder is not None:
            unitHolder = specifiedUnitHolder
        else:
            unitHolder = self._conceptToUnitHolderMap.get(concept)

        if unitHolder:
            cell = self.getSingleCell(unitHolder, row=row)
            if cell is None or cell.value is None:
                self._results.addMessage(
                    f"Unable to find unit in expected part of {unitHolder.definedName.name}. Related concept {conceptHolder.definedName.name} has coordinates {excelCellRangeRef(conceptHolder.worksheet, conceptHolder.cellRange)}.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(
                        unitHolder.worksheet, unitHolder.cellRange
                    ),
                )
                return False
            if (unit := self.getSimpleUnit(unitHolder, cell)) is not None:
                if self.taxonomy.UTR.valid(concept.dataType, unit):
                    factBuilder.setSimpleUnit(unit)
                    return True
                elif specifiedUnitHolder:
                    if not sharedRange:
                        self._results.addMessage(
                            f"Unable to create fact due to specified cell value '{cell.value}' not matching data type '{concept.dataType}'.",
                            Severity.WARNING,
                            MessageType.Conversion,
                            taxonomy_concept=concept,
                            excel_reference=excelCellRef(unitHolder.worksheet, cell),
                        )
                    return False
                else:
                    self._results.addMessage(
                        f"Found unit {unit} for {unitHolder.definedName.name} but it is not valid for {concept.qname} with dataType {concept.dataType}. Attempting fallback unit. Cell value '{cell.value}'.",
                        Severity.ERROR,
                        MessageType.DevInfo,
                        excel_reference=excelCellRef(unitHolder.worksheet, cell),
                    )
                    return self.setFallbackUnitForName(
                        conceptHolder.definedName, concept, factBuilder
                    )
            # If units have really broken _unit ranges or dodgy labels, we might
            # have just set a default unit for the concept.
            elif (unitQname := self._configConceptToUnitMap.get(concept)) is not None:
                if self.taxonomy.UTR.valid(concept.dataType, unitQname):
                    self._results.addMessage(
                        f"Using configured unit {unitQname} for {concept} as unit cell value could not be translated in to a unit. Cell value '{cell.value}'.",
                        Severity.ERROR,
                        MessageType.DevInfo,
                        excel_reference=excelCellRef(unitHolder.worksheet, cell),
                    )
                    factBuilder.setSimpleUnit(unitQname)
                    return True
                else:
                    self._results.addMessage(
                        f"Unit override in config is broken. Unit {unitQname} is not valid for {concept} with dataType {concept.dataType}.",
                        Severity.ERROR,
                        MessageType.DevInfo,
                        excel_reference=excelCellRangeRef(
                            conceptHolder.worksheet, conceptHolder.cellRange
                        ),
                    )
            else:
                self._results.addMessage(
                    f"Unable to find unit for {unitHolder.definedName.name} using named range. Attempting to find unit via taxonomy. Cell value '{cell.value}'.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                    excel_reference=excelCellRef(unitHolder.worksheet, cell),
                )

        # simple-units: Template is out of ideas for a unit, try the measurement guidance label
        if (units := concept.getRequiredUnitQNames()) is not None:
            if 1 == len(units):
                factBuilder.setSimpleUnit(next(iter(units)))
                return True
            else:
                self._results.addMessage(
                    f"No unit found in Excel for {conceptHolder.definedName.name}. More than one unit specified as possible in the taxonomy. {units=}",
                    Severity.WARNING,
                    MessageType.Conversion,
                    taxonomy_concept=concept,
                    excel_reference=excelCellRangeRef(
                        conceptHolder.worksheet, conceptHolder.cellRange
                    ),
                )
                return False

        # complex-units: Template is out of ideas for a unit, see if we have a complex unit in the configuration.
        candidateUnitIds = list(
            self.taxonomy.UTR.getUnitIdsForDataType(concept.dataType)
        )
        for c in candidateUnitIds:
            complex_unit = self._configUnitIdsToMeasures.get(c)
            if complex_unit is not None:
                factBuilder.setComplexUnit(
                    complex_unit.numerator, complex_unit.denominator
                )
                return True

        return self.setFallbackUnitForName(
            conceptHolder.definedName, concept, factBuilder
        )

    def createSimpleFacts(self) -> None:
        reportable = {
            dn: stuff
            for dn, stuff in self._definedNameToXBRLMap.items()
            if (c := stuff.concept) and c.isReportable
        }

        for dn, stuff in reportable.copy().items():
            required_dims = self.taxonomy.getExplicitDimensionsForPrimaryItem(
                stuff.concept
            )
            preset_dims = frozenset(self._presetDimensions.get(stuff, {}).keys())
            unset_dims = required_dims.difference(
                self.taxonomy.defaultedDimensions, preset_dims
            )
            if unset_dims:
                self._results.addMessage(
                    f"The named range {dn.name} has required dimensions that have not been set.\n The required dimensions {conceptsToText(required_dims)}.\n Missing: {conceptsToText(unset_dims)}.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                reportable.pop(dn)

        for dn, stuff in reportable.items():
            concept = stuff.concept
            assert concept.isReportable

            fb = self._report.getFactBuilder()

            if concept.isEnumerationSet:
                self.createEESetFact(stuff, fb)
                self._definedNameToXBRLMap.pop(dn)
                continue

            cell = self.getSingleCell(dn)
            if cell is None:
                self._definedNameToXBRLMap.pop(dn)
                continue
            value = cell.value
            if value is None or value is False:
                # No value in the cell, so not reportable
                self._definedNameToXBRLMap.pop(dn)
                continue
            if value in EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE:
                # Placeholder values are not reportable
                self._definedNameToXBRLMap.pop(dn)
                continue

            if concept.isDate:
                try:
                    value = self.getDateFromValue(value)
                except Exception:
                    self._results.addMessage(
                        f"Unable to parse date from cell value '{value}' for {concept.qname}.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        taxonomy_concept=concept,
                        excel_reference=excelCellRef(stuff.worksheet, cell),
                    )
                    self._definedNameToXBRLMap.pop(dn)
                    continue

            fb.setConcept(concept).setValue(value)
            if concept.isNumeric:
                self.processNumeric(stuff, cell, fb, value)
            if concept.isNumeric and not concept.isMonetary:
                self.setUnitForName(stuff, fb)
            elif concept.isMonetary:
                pass  # monetary are all assumed to be the same currency (set via defaults)
            elif concept.isEnumerationSingle:
                s_value = str(value)
                eeValue = self._report.taxonomy.getConceptForLabel(s_value)
                warn = False
                if (
                    eeValue is None
                    and (
                        fake_value := self._configCellValuesToTaxonomyLabels.get(
                            s_value
                        )
                    )
                    is not None
                ):
                    eeValue = self._report.taxonomy.getConceptForLabel(fake_value)
                    warn = True
                if eeValue is not None:
                    fb.setHiddenValue(eeValue.expandedName)
                    if warn:
                        self._results.addMessage(
                            f"Workaround performed for EE member label mismatch when reporting {concept.qname}. Cell value '{value}'. Concept label '{eeValue.getStandardLabel()}'",
                            Severity.WARNING,
                            MessageType.DevInfo,
                            taxonomy_concept=concept,
                        )
                else:
                    eeDomainLabels: dict[str, Concept] = dict(
                        (label.replace("[member]", "").strip(), member)
                        for member in concept.getEEDomain()
                        if (label := member.getStandardLabel()) is not None
                    )
                    closest_matches: Optional[list[str]] = difflib.get_close_matches(
                        s_value, eeDomainLabels.keys(), n=1, cutoff=0.6
                    )
                    if closest_matches:
                        eeValue = eeDomainLabels[closest_matches[0]]
                        fb.setHiddenValue(eeValue.expandedName)
                        self._results.addMessage(
                            f"Using closest match EE concept when reporting {concept.qname}. Cell value '{value}'. Chosen EE domain member: {eeValue.qname}; part of label used: '{closest_matches[0]}'",
                            Severity.WARNING,
                            MessageType.Conversion,
                        )
                    else:
                        self._results.addMessage(
                            f"Unable to find EE concept when reporting {concept.qname}. Cell value '{value}'. EE domain: {eeDomainAsText(concept)}",
                            Severity.ERROR,
                            MessageType.Conversion,
                        )

            if (presetDimensions := self._presetDimensions.get(stuff)) is not None:
                for dim, dimValue in presetDimensions.items():
                    defaultValue = self.taxonomy.getDimensionDefault(dim)
                    if defaultValue is None or dimValue != defaultValue:
                        fb.setExplicitDimension(dim, dimValue)

                    dimValueDN: Optional[DefinedName] = None
                    if (
                        dimValueDN := self._workbook.defined_names.get(
                            dimValue.qname.localName
                        )
                    ) is None:
                        continue

                    namedPeriod: str = dimValueDN.name
                    if self._report.hasNamedPeriod(namedPeriod):
                        fb.setNamedPeriod(namedPeriod)

            self._definedNameToXBRLMap.pop(dn)
            self.addFactToReport(fb, stuff)

    def createEESetFact(
        self, stuff: CellAndXBRLMetadataHolder, fb: FactBuilder
    ) -> None:
        concept = stuff.concept
        assert concept.isEnumerationSet
        eeSetValue: set[Concept] = set()
        value: list[str] = []
        eeDomain = concept.getEEDomain()

        for rnum, cnum, cell in getCellRangeIterator(stuff.worksheet, stuff.cellRange):
            v = cell.value
            if v is None or v is False:
                continue
            if v is True:
                rindex = rnum - stuff.cellRange.min_row
                cindex = cnum - stuff.cellRange.min_col
                if 1 == stuff.effectiveHeight:
                    index = cindex
                elif 1 == stuff.effectiveWidth:
                    index = rindex
                elif stuff.effectiveHeight < stuff.effectiveWidth:
                    index = cindex
                else:
                    index = rindex

                if 0 <= index < len(eeDomain):
                    eeMember = eeDomain[index]
                else:
                    self._results.addMessage(
                        "Failed to process enumeration value",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        taxonomy_concept=stuff.concept,
                        excel_reference=excelCellRef(stuff.worksheet, cell),
                    )
                    L.error(
                        f"Trying to access cell in named range {stuff.definedName.name} {rnum=} {cnum=} {stuff.cellRange.bounds=} {index=} {len(eeDomain)}"
                    )
                    continue
                eeSetValue.add(eeMember)
                value.append(
                    eeMember.getStandardLabel(
                        fallbackIfMissing=str(eeMember.qname), removeSuffix=True
                    )
                )
            elif isinstance(v, str) and v == EE_SET_DESIRED_EMPTY_PLACEHOLDER_VALUE:
                value.append(v)
            elif isinstance(v, str):
                warn = False
                e_label = v
                if v.startswith("NACE "):
                    e_label = v.replace("NACE ", "")
                    warn = True
                eeConcept = self._report.taxonomy.getConceptForLabel(e_label)
                if (
                    eeConcept is None
                    and (
                        fake_value := self._configCellValuesToTaxonomyLabels.get(
                            e_label
                        )
                    )
                    is not None
                ):
                    warn = True
                    eeConcept = self._report.taxonomy.getConceptForLabel(fake_value)
                if eeConcept is not None:
                    value.append(str(v))
                    eeSetValue.add(eeConcept)
                    if warn:
                        self._results.addMessage(
                            f"Workaround performed for EE member label mismatch when reporting {concept.qname}. Cell value '{v}'. Concept label '{eeConcept.getStandardLabel()}'",
                            Severity.WARNING,
                            MessageType.DevInfo,
                            taxonomy_concept=concept,
                            excel_reference=excelCellRef(stuff.worksheet, cell),
                        )
                else:
                    self._results.addMessage(
                        f"Unable to find EE member when reporting {concept.qname}. Cell value '{v}'.",
                        Severity.ERROR,
                        MessageType.ExcelParsing,
                        taxonomy_concept=concept,
                        excel_reference=excelCellRef(stuff.worksheet, cell),
                    )
            else:
                self._results.addMessage(
                    f"Unable to find EE domain member when reporting {concept.qname}. Cell value '{v}'",
                    Severity.ERROR,
                    MessageType.Conversion,
                    taxonomy_concept=concept,
                    excel_reference=excelCellRef(stuff.worksheet, cell),
                )
        if EE_SET_DESIRED_EMPTY_PLACEHOLDER_VALUE in value:
            onlyPlaceholder = set([EE_SET_DESIRED_EMPTY_PLACEHOLDER_VALUE])
            otherValues = set(x for x in value if x is not None).difference(
                onlyPlaceholder
            )
            if otherValues:
                self._results.addMessage(
                    f"Inconsistent values found for EE set {concept.qname}. Not creating an XBRL fact. Cell values '{value}'",
                    Severity.ERROR,
                    MessageType.Conversion,
                    taxonomy_concept=concept,
                    excel_reference=excelCellRangeRef(stuff.worksheet, stuff.cellRange),
                )
            else:
                # Mechanism to say "I want to create an empty EE set fact"
                fb.setConcept(concept).setHiddenValue("").setValue(
                    EE_SET_DESIRED_EMPTY_PLACEHOLDER_VALUE
                )
                self.addFactToReport(fb, stuff)
        elif not eeSetValue:
            self._results.addMessage(
                f"No values found for {concept.qname} so not creating an empty XBRL fact. Cell value '{value}'",
                Severity.INFO,
                MessageType.DevInfo,
                taxonomy_concept=concept,
                excel_reference=excelCellRef(stuff.worksheet, cell),
            )
        else:
            fb.setConcept(concept).setHiddenValue(
                " ".join(sorted(e.expandedName for e in eeSetValue))
            ).setValue("\n".join(value))
            self.addFactToReport(fb, stuff)
        return None

    def setFallbackUnitForName(
        self, dn: DefinedName, concept: Concept, factBuilder: FactBuilder
    ) -> bool:
        if not concept.isNumeric:
            # can't set unit!
            return False

        # If we have a default unit for the data type, use it iff UTR valid.
        if (unit := self._configDataTypeToUnitMap.get(concept.dataType)) is not None:
            if self.taxonomy.UTR.valid(concept.dataType, unit):
                factBuilder.setSimpleUnit(unit)
                return True

        # Otherwise pick the first unit from the UTR that is valid.
        if units := self.taxonomy.UTR.getUnitsForDataType(concept.dataType):
            chosen = next(iter(units))
            self._results.addMessage(
                f"Picked fallback unit (from UTR) {chosen} for {dn.name}",
                Severity.WARNING,
                MessageType.DevInfo,
            )
            factBuilder.setSimpleUnit(chosen)
        else:
            # At this point we know that xbrli:pure is guaranteed to exist and
            # that the data type is not in the UTR so xbrli:pure won't cause a
            # UTR validation failure.
            # TODO: Think about complex units.
            ultimateFallback = self.taxonomy.QNameMaker.fromString("xbrli:pure")
            self._results.addMessage(
                f"Used ultimate fallback unit {ultimateFallback} for {dn.name}",
                Severity.WARNING,
                MessageType.DevInfo,
            )
            factBuilder.setSimpleUnit(ultimateFallback)
        return True

    def processNumeric(
        self,
        stuff: CellAndXBRLMetadataHolder,
        cell: Cell,
        fb: FactBuilder,
        value: Optional[int] = None,
    ) -> None:
        if value is None:
            value = cell.value
        decimals = get_decimal_places(cell)

        cell_is_percentage = "%" in cell.number_format
        if fb._concept is not None:
            concept_is_percentage = "percentItemType" == fb._concept.dataType.localName
            if cell_is_percentage != concept_is_percentage:
                self._results.addMessage(
                    f"Cell number format and XBRL Taxonomy data type disagree about percentages. Cell number format '{cell.number_format}'. Concept data type {fb._concept.dataType}.",
                    Severity.WARNING,
                    MessageType.DevInfo,
                    taxonomy_concept=fb._concept,
                    excel_reference=excelCellRef(stuff.worksheet, cell),
                )

        if cell_is_percentage:
            fb.setPercentageValue(value, decimals, inputIsDecimalForm=True)
        else:
            fb.setDecimals(decimals)
        return

    def getOrAddNamedPeriodForYear(
        self,
        name: str,
        year: int,
    ) -> str:
        """
        Given a year and optional start/end date templates, return named period for that year.
        The day and month will be taken from the templates, and the year will be set to `year`.
        If no template is provided, the day and month for the start and end default to Jan 1 and Dec 31.
        """
        if self._report.hasNamedPeriod(name):
            return name
        # Need to add a new named period for the year.
        endOfDefault = self._report.defaultPeriod.end
        end = endOfDefault + relativedelta(year=year)
        start = end + relativedelta(years=-1, days=+1)
        self._report.addDurationPeriod(name, start, end)
        return name

    def checkForUnhandledItems(self) -> None:
        """
        Check for any items that have accidentally not been turned in to XBRL facts.
        """
        unHandled: list[CellAndXBRLMetadataHolder] = []
        unHandled.extend(self._definedNameToXBRLMap.values())
        # TODO: mark units as used once they have been used, then we can check here.
        # unHandled.extend(self._conceptToUnitHolderMap.values())

        # FIXME: temporary workaround for VSME taxonomy.
        ignore_dns = {"BreakdownOfEnergyConsumptionAxis"}
        # FIXME: temporary workaround for VSME taxonomy.

        for stuff in unHandled:
            if stuff.definedName.name in ignore_dns:
                continue
            message = f"Failed to handle XBRL related Excel named range {stuff.definedName.name}."
            self._results.addMessage(
                message,
                Severity.ERROR,
                MessageType.Conversion,
            )

    def getSingleDateValue(self, definedName: DefinedName | str) -> date:
        value = self.getSingleValue(definedName)
        return self.getDateFromValue(value)

    def getDateFromValue(self, value: _CellValue) -> date:
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, date):
            return value
        elif isinstance(value, str):
            if "-" in value:
                return date.fromisoformat(value)
            elif "/" in value:
                return parse_datetime(value, yearfirst=False, dayfirst=True).date()
            raise ValueError(f"Unsupported date string: '{value}'")
        else:
            raise TypeError(
                f"Unsupported type for date conversion: {type(value).__name__}"
            )
